"""
============================================================
  Canada IRCC — Non-Compliant Employers Scraper
  Source: https://www.canada.ca/en/immigration-refugees-
          citizenship/services/work-canada/employers-non-compliant.html
============================================================
  Exports data to:
    • non_compliant_employers.xlsx  (formatted Excel)
    • non_compliant_employers.csv   (plain CSV)

  Run:  python scrape_non_compliant.py
============================================================
"""

import time
import sys
from datetime import datetime
from pathlib import Path

# ── Dependency check ─────────────────────────────────────
def check_deps():
    missing = []
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        missing.append("playwright")
    try:
        import pandas
    except ImportError:
        missing.append("pandas")
    try:
        import openpyxl
    except ImportError:
        missing.append("openpyxl")
    if missing:
        print(f"\n❌  Missing packages: {', '.join(missing)}")
        print(f"    Run:  pip install {' '.join(missing)}")
        if "playwright" in missing:
            print("    Then: playwright install chromium")
        sys.exit(1)

check_deps()

from playwright.sync_api import sync_playwright
import pandas as pd
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── Config ────────────────────────────────────────────────
URL = (
    "https://www.canada.ca/en/immigration-refugees-citizenship"
    "/services/work-canada/employers-non-compliant.html"
)

EXPECTED_COLUMNS = [
    "Business Operating Name",
    "Business Legal Name",
    "Address",
    "Reason(s)",
    "Date of Final Decision",
    "Penalty",
    "Status",
]

OUTPUT_DIR = Path(__file__).parent
TIMESTAMP  = datetime.now().strftime("%Y-%m-%d_%H-%M")
XLSX_FILE  = OUTPUT_DIR / f"non_compliant_employers_{TIMESTAMP}.xlsx"
CSV_FILE   = OUTPUT_DIR / f"non_compliant_employers_{TIMESTAMP}.csv"


# ── Helpers ───────────────────────────────────────────────
def wait_for_table(page):
    """Wait until the table has at least one populated data row."""
    try:
        page.wait_for_function(
            """() => {
                const rows = document.querySelectorAll('table tbody tr');
                return rows.length > 0 &&
                       rows[0].querySelectorAll('td')[0]?.innerText.trim().length > 0;
            }""",
            timeout=20_000,
        )
    except Exception:
        time.sleep(3)  # fallback wait


def extract_page_rows(page):
    """Pull all data rows from the current page view."""
    return page.eval_on_selector_all(
        "table tbody tr",
        """rows => rows.map(row =>
            Array.from(row.querySelectorAll('td'))
                .map(td => td.innerText.trim())
        )"""
    )


def find_next_button(page):
    """
    Try several selectors to find the Next page button.
    Returns the element handle or None.
    """
    selectors = [
        "a[aria-label='Next page']",
        "a[aria-label='next page']",
        "button[aria-label='Next page']",
        "li.next a",
        "a.next",
        "[class*='next'] a",
        "[class*='pagination'] a[rel='next']",
        "a:has-text('Next')",
        "button:has-text('Next')",
    ]
    for sel in selectors:
        try:
            el = page.query_selector(sel)
            if el and el.is_visible():
                return el
        except Exception:
            continue
    return None


def is_next_disabled(page):
    """Return True if the Next button exists but is disabled/greyed out."""
    disabled_selectors = [
        "li.next.disabled",
        "li[class*='next'][class*='disabled']",
        "a[aria-label='Next page'][aria-disabled='true']",
        ".pagination .next.disabled",
    ]
    for sel in disabled_selectors:
        try:
            el = page.query_selector(sel)
            if el:
                return True
        except Exception:
            continue
    return False


# ── Scraper ───────────────────────────────────────────────
def scrape():
    print("\n🌐  Launching browser...")
    all_rows = []
    headers  = EXPECTED_COLUMNS

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True)
        context = browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/122.0.0.0 Safari/537.36"
            )
        )
        page = context.new_page()

        print(f"📄  Loading page 1...")
        page.goto(URL, wait_until="networkidle", timeout=60_000)
        wait_for_table(page)

        # ── Grab headers from first page ──────────────────
        detected = page.eval_on_selector_all(
            "table thead th",
            "els => els.map(e => e.innerText.trim()).filter(t => t.length > 0)"
        )
        if detected:
            headers = detected
        print(f"📋  Columns: {headers}")

        current_page = 1

        while True:
            # Extract rows on current page
            rows_raw = extract_page_rows(page)
            rows = [r for r in rows_raw if any(cell for cell in r)]
            all_rows.extend(rows)
            print(f"   Page {current_page:>2}  →  {len(rows):>3} records  "
                  f"(total so far: {len(all_rows)})")

            # Check if Next is disabled (last page)
            if is_next_disabled(page):
                print("   ↳ Last page reached (Next is disabled).")
                break

            # Find and click the Next button
            next_btn = find_next_button(page)
            if not next_btn:
                print("   ↳ No Next button found — assuming last page.")
                break

            # Click Next and wait for new table data
            next_btn.click()
            time.sleep(1.5)
            wait_for_table(page)
            time.sleep(0.5)
            current_page += 1

        browser.close()

    print(f"\n✅  {len(all_rows)} total records across {current_page} pages.")

    # ── Build list of dicts ───────────────────────────────
    records = []
    for row in all_rows:
        while len(row) < len(headers):
            row.append("")
        record = {headers[i]: row[i] for i in range(len(headers))}
        records.append(record)

    return records, headers


# ── Excel export ──────────────────────────────────────────
def to_excel(records: list[dict], headers: list[str]):
    df = pd.DataFrame(records, columns=headers)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Non-Compliant Employers"

    # ── Palette ───────────────────────────────────────────
    RED_DARK   = "C0392B"
    RED_LIGHT  = "FADBD8"
    GREY_HEAD  = "2C3E50"
    WHITE      = "FFFFFF"
    STRIPE     = "F9EBEA"
    BORDER_CLR = "BDC3C7"

    thin = Side(style="thin", color=BORDER_CLR)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Title row ─────────────────────────────────────────
    ws.merge_cells("A1:G1")
    title_cell = ws["A1"]
    title_cell.value = "🇨🇦  IRCC — Employers Found Non-Compliant"
    title_cell.font = Font(name="Calibri", bold=True, size=14, color=WHITE)
    title_cell.fill = PatternFill("solid", fgColor=RED_DARK)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # ── Subtitle / timestamp ──────────────────────────────
    ws.merge_cells("A2:G2")
    sub = ws["A2"]
    sub.value = (
        f"Downloaded: {datetime.now().strftime('%B %d, %Y at %H:%M')}  |  "
        f"Source: canada.ca/en/immigration-refugees-citizenship"
        f"/services/work-canada/employers-non-compliant.html  |  "
        f"Total records: {len(records)}"
    )
    sub.font = Font(name="Calibri", italic=True, size=9, color="7F8C8D")
    sub.fill = PatternFill("solid", fgColor="FDFEFE")
    sub.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 18

    # ── Column headers ────────────────────────────────────
    for col_idx, col_name in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=col_name)
        cell.font      = Font(name="Calibri", bold=True, size=10, color=WHITE)
        cell.fill      = PatternFill("solid", fgColor=GREY_HEAD)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border
    ws.row_dimensions[3].height = 24

    # ── Data rows ─────────────────────────────────────────
    for row_idx, record in enumerate(records, start=4):
        is_stripe = (row_idx % 2 == 0)
        bg = PatternFill("solid", fgColor=STRIPE if is_stripe else WHITE)

        for col_idx, col_name in enumerate(headers, start=1):
            val  = record.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = Font(name="Calibri", size=9)
            cell.fill      = bg
            cell.border    = border
            cell.alignment = Alignment(
                vertical="top", wrap_text=True,
                horizontal="center" if col_idx > 2 else "left"
            )

            # Highlight banned / non-compliant status
            if col_name.lower() == "status" and val:
                cell.font = Font(name="Calibri", size=9, bold=True, color=RED_DARK)
                cell.fill = PatternFill("solid", fgColor=RED_LIGHT)

        ws.row_dimensions[row_idx].height = 40

    # ── Column widths ─────────────────────────────────────
    col_widths = {1: 28, 2: 28, 3: 30, 4: 18, 5: 18, 6: 16, 7: 16}
    for col_idx, width in col_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Freeze panes & filters ────────────────────────────
    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(len(headers))}{len(records)+3}"

    # ── Summary sheet ─────────────────────────────────────
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Summary"
    ws2["A1"].font = Font(bold=True, size=13, color=RED_DARK)
    ws2["A2"] = f"Total non-compliant employers:  {len(records)}"
    ws2["A3"] = f"Downloaded on:  {datetime.now().strftime('%B %d, %Y')}"
    ws2["A4"] = f"Source URL:  {URL}"
    ws2["A5"].value = ""
    ws2["A6"] = "Status Breakdown"
    ws2["A6"].font = Font(bold=True)

    if records:
        status_col = headers.index("Status") if "Status" in headers else -1
        if status_col >= 0:
            from collections import Counter
            statuses = Counter(r.get("Status", "Unknown") for r in records)
            for i, (status, count) in enumerate(statuses.most_common(), start=7):
                ws2[f"A{i}"] = status
                ws2[f"B{i}"] = count

    for col in ["A", "B"]:
        ws2.column_dimensions[col].width = 40

    wb.save(XLSX_FILE)
    print(f"📊  Excel saved  →  {XLSX_FILE.name}")


# ── CSV export ────────────────────────────────────────────
def to_csv(records: list[dict], headers: list[str]):
    df = pd.DataFrame(records, columns=headers)
    df.to_csv(CSV_FILE, index=False, encoding="utf-8-sig")  # utf-8-sig for Excel compat
    print(f"📄  CSV saved    →  {CSV_FILE.name}")


# ── Main ──────────────────────────────────────────────────
def main():
    print("=" * 58)
    print("  IRCC Non-Compliant Employers Scraper")
    print("=" * 58)

    records, headers = scrape()

    if not records:
        print("\n⚠️   No records found. The page may have changed or")
        print("    data may not have loaded. Try running again.")
        sys.exit(1)

    to_excel(records, headers)
    to_csv(records, headers)

    print("\n✅  Done!")
    print(f"    Files saved to: {OUTPUT_DIR.resolve()}")
    print("=" * 58)


if __name__ == "__main__":
    main()
